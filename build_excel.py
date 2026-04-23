"""Genera catalogo.xlsx con 90 perfumes — seed para importar a Google Sheets.

Imágenes: CDN de Fragrantica → https://fimgs.net/mdimg/perfume/375x500.{id}.jpg
Los que no tienen ID confirmado quedan sin imagen (el HTML muestra SVG placeholder).
Precios benchmark referencia Farmashop Uruguay abr 2026.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "Catalogo"

AZUL = "0B3D91"
AZUL_CLARO = "E8F0FE"
BLANCO = "FFFFFF"
GRIS = "6B7280"

headers = [
    "ID", "Marca", "Nombre", "Genero", "Tamaño (ml)",
    "Notas / Descripcion", "Precio (UYU)", "Precio antes (UYU)",
    "Stock", "Promocion", "Imagen URL"
]

def IMG(fid): return f"https://fimgs.net/mdimg/perfume/375x500.{fid}.jpg" if fid else ""

# =========================================================
#  CATÁLOGO DE 90 PERFUMES
#  Estructura: (ID, Marca, Nombre, Género, ml, Notas, Precio, PrecioAntes, Stock, Promocion, Imagen)
# =========================================================
data = [
    # ───── HOMBRE (1-50) ─────
    (1,  "Dior",              "Sauvage EDT",                    "Hombre", 100, "Bergamota, pimienta de Sichuán, ambroxan. Fresco y especiado.",          10600, None,  "SI", "NO", IMG(31861)),
    (2,  "Paco Rabanne",      "One Million EDT",                "Hombre", 100, "Cuero, canela, pomelo sangriento. Dulce y magnético.",                     4590, 5100,  "SI", "SI", IMG(3747)),
    (3,  "Carolina Herrera",  "212 VIP Men NYC EDT",            "Hombre", 100, "Gin tónic, menta, cuero vegetal. Fresco, nocturno.",                       5800, None,  "SI", "NO", IMG(12865)),
    (4,  "Giorgio Armani",    "Acqua di Giò EDT",               "Hombre", 100, "Marino, bergamota, pachulí. Fresco acuático icónico.",                     7500, None,  "SI", "NO", IMG(410)),
    (5,  "Paco Rabanne",      "Invictus EDT",                   "Hombre", 100, "Pomelo, laurel marino, ambergris. Deportivo y magnético.",                 5200, None,  "SI", "NO", IMG(18471)),
    (11, "Dior",              "Sauvage EDP",                    "Hombre", 100, "Bergamota, pimienta, lavanda, vainilla ambarada. Más profundo que el EDT.", 12500, None, "SI", "NO", IMG(48100)),
    (12, "Dior",              "Homme Intense 2011",             "Hombre", 100, "Iris polvoriento, lavanda, ámbar, pera. Masculino, elegante, invernal.",   12800, None, "SI", "NO", IMG(13016)),
    (13, "Dior",              "Homme EDT",                      "Hombre", 100, "Iris, salvia, vetiver. Fresco, refinado y misterioso.",                    12000, None, "SI", "NO", IMG(13015)),
    (14, "Dior",              "Fahrenheit EDT",                 "Hombre", 100, "Espino, madreselva, violeta, cuero. Clásico atemporal.",                    9200, None, "SI", "NO", IMG(384)),
    (15, "Dior",              "Eau Sauvage EDT",                "Hombre", 100, "Limón, albahaca, romero. Cítrico aromático clásico.",                       8400, None, "SI", "NO", IMG(231)),
    (16, "Paco Rabanne",      "Phantom EDT",                    "Hombre", 100, "Lavanda, limón, cardamomo, vetiver. Fresco tecno-moderno.",                 5900, None, "SI", "NO", IMG(68226)),
    (17, "Paco Rabanne",      "Pure XS EDT",                    "Hombre", 100, "Tomillo, jengibre, canela, cuero. Oriental especiado sensual.",             5400, None, "SI", "NO", IMG(46038)),
    (18, "Paco Rabanne",      "1 Million Parfum",               "Hombre", 100, "Mandarina sanguina, cardamomo, cuero ambarado. Más intenso que el EDT.",    5800, None, "SI", "NO", IMG(60035)),
    (19, "Paco Rabanne",      "Invictus Victory Elixir",        "Hombre", 100, "Olíbano, caoba, pachulí. Seductor, oriental especiado.",                    6200, None, "SI", "NO", IMG(78575)),
    (20, "Paco Rabanne",      "Invictus Parfum",                "Hombre", 100, "Laurel, jazmín, ambroxan. Acuático aromático intensificado.",               5900, None, "SI", "NO", IMG(90433)),
    (21, "Carolina Herrera",  "212 Men EDT",                    "Hombre", 100, "Bergamota, pimienta verde, incienso, gengibre. Fresco urbano.",             5400, None, "SI", "NO", IMG(297)),
    (22, "Carolina Herrera",  "212 VIP Black EDP",              "Hombre", 100, "Menta, ron, cuero, regaliz. Intenso, nocturno, dulce.",                     6200, None, "SI", "NO", IMG(46093)),
    (23, "Carolina Herrera",  "Bad Boy EDT",                    "Hombre", 100, "Pimienta blanca, salvia, cacao, tonka. Amargo gourmand moderno.",           6800, None, "SI", "NO", IMG(55449)),
    (24, "Carolina Herrera",  "Bad Boy Cobalt EDP",             "Hombre", 100, "Pimienta, sabio, plum, cedro, vainilla. Más frutal e intenso.",             7500, None, "SI", "NO", IMG(71888)),
    (25, "Carolina Herrera",  "CH Men EDT",                     "Hombre", 100, "Pomelo, lavanda, canela, cuero. Elegante, sofisticado.",                    5600, None, "SI", "NO", ""),
    (26, "Giorgio Armani",    "Acqua di Giò Profumo",           "Hombre", 100, "Incienso, salvia, menta, pachulí. Versión más intensa del Giò.",            9800, None, "SI", "NO", IMG(29727)),
    (27, "Giorgio Armani",    "Acqua di Giò Parfum",            "Hombre", 100, "Vetiver, pachulí, incienso. Intensísimo, recargable.",                     11200, None, "SI", "NO", IMG(81508)),
    (28, "Giorgio Armani",    "Armani Code EDT",                "Hombre",  75, "Anís estrellado, limón, oliva, tonka. Oriental aromático dulce.",           7800, None, "SI", "NO", IMG(412)),
    (29, "Giorgio Armani",    "Stronger With You EDT",          "Hombre", 100, "Castaña, menta, salvia, vainilla, smoke. Dulce gourmand moderno.",          6800, None, "SI", "NO", IMG(45258)),
    (30, "Hugo Boss",         "Boss Bottled EDT",               "Hombre", 100, "Manzana, canela, clavo, sándalo. Clásico masculino universal.",             5200, None, "SI", "NO", IMG(383)),
    (31, "Hugo Boss",         "Boss Bottled Parfum",            "Hombre", 100, "Canela, manzana, cuero, cedro. Intensificado, más persistente.",            7500, None, "SI", "NO", IMG(75183)),
    (32, "Hugo Boss",         "Boss The Scent EDT",             "Hombre", 100, "Mandarina, maninka, cuero. Seductor especiado.",                            5800, None, "SI", "NO", IMG(31445)),
    (33, "Hugo Boss",         "Hugo EDT",                   "Hombre", 125, "Menta, manzana verde, lavanda, abeto. Fresco juvenil icónico.",             4200, None, "SI", "NO", IMG(570)),
    (34, "Lacoste",           "L.12.12 Blanc EDT",              "Hombre", 100, "Pomelo, cardamomo, romero, cedro. Limpio y fresco.",                        4800, None, "SI", "NO", IMG(11043)),
    (35, "Lacoste",           "Essential EDT",                  "Hombre", 125, "Uva, pepino, vetiver, cedro. Verde fresco deportivo.",                      4500, None, "SI", "NO", IMG(787)),
    (36, "Calvin Klein",      "Eternity Men EDT",               "Hombre", 100, "Lavanda, geranio, cedro, sándalo. Aromático clásico noventero.",            5200, None, "SI", "NO", IMG(258)),
    (37, "Calvin Klein",      "CK One EDT",                     "Unisex", 200, "Limón, bergamota, papaya, té verde. Unisex fresco icónico.",                4200, None, "SI", "NO", IMG(276)),
    (38, "Calvin Klein",      "Obsession for Men EDT",          "Hombre", 125, "Mandarina, lavanda, amberes, almizcle. Oriental especiado 80s.",           4600, None, "SI", "NO", IMG(251)),
    (39, "Versace",           "Eros EDT",                       "Hombre", 100, "Menta, manzana verde, tonka, vainilla, vetiver. Dulce aromático icónico.",  6500, None, "SI", "NO", IMG(16657)),
    (40, "Versace",           "Eros EDP",                       "Hombre", 100, "Menta, manzana, madera, vainilla cremosa. Más oriental que el EDT.",        7200, None, "SI", "NO", IMG(62762)),
    (41, "Versace",           "Pour Homme Dylan Blue EDT",      "Hombre", 100, "Bergamota, pomelo, pimienta rosa, pachulí. Moderno fresco.",                6200, None, "SI", "NO", IMG(40031)),
    (42, "Versace",           "Pour Homme EDT",                 "Hombre", 100, "Bergamota, neroli, clary sage, cedro. Fresco aromático mediterráneo.",      5800, None, "SI", "NO", IMG(2318)),
    (43, "Azzaro",            "Wanted EDT",                     "Hombre", 100, "Limón, pimienta, cardamomo, cuero, vetiver. Especiado carismático.",        5800, None, "SI", "NO", IMG(38686)),
    (44, "Azzaro",            "Chrome EDT",                     "Hombre", 100, "Limón, bergamota, jazmín, almizcle. Fresco acuático.",                      5200, None, "SI", "NO", IMG(788)),
    (45, "Tommy Hilfiger",    "Tommy EDT",                      "Hombre", 100, "Lavanda, manzana, menta, cuero. Fresco americano casual.",                  4500, None, "SI", "NO", IMG(6314)),
    (46, "Nautica",           "Voyage EDT",                     "Hombre", 100, "Manzana, loto, abeto, musgo. Acuático aromático fresco.",                   3800, None, "SI", "NO", IMG(913)),
    (47, "Jean Paul Gaultier","Le Male EDT",                    "Hombre", 125, "Lavanda, menta, cardamomo, vainilla tonka. Oriental aromático icónico.",    6500, None, "SI", "NO", IMG(430)),
    (48, "Jean Paul Gaultier","Ultra Male EDT",                 "Hombre", 125, "Pera negra, lavanda, canela, vainilla ámbar. Más dulce e intenso.",         7200, None, "SI", "NO", IMG(30947)),
    (49, "Yves Saint Laurent","Y EDP",                          "Hombre", 100, "Manzana, jengibre, salvia, ambroxan. Fresco amaderado moderno.",            8200, None, "SI", "NO", IMG(50757)),
    (50, "Yves Saint Laurent","L'Homme EDT",                    "Hombre", 100, "Jengibre, bergamota, cedro, tonka. Elegante masculino.",                    7800, None, "SI", "NO", IMG(734)),

    # ───── MUJER (6-10 y 51-90) ─────
    (6,  "Carolina Herrera",  "Good Girl EDP",                  "Mujer",   80, "Jazmín, almendra, cacao, tonka. Floral dulce elegante.",                    6800, None,  "SI", "NO", IMG(39681)),
    (7,  "Lancôme",           "La Vie Est Belle EDP",           "Mujer",  100, "Iris, pera, pachulí, vainilla. Gourmand luminoso.",                         8200, None,  "SI", "NO", IMG(14982)),
    (8,  "Chanel",            "Coco Mademoiselle EDP",          "Mujer",  100, "Rosa, jazmín, pachulí, vetiver. Oriental chic atemporal.",                 10500, None,  "SI", "NO", IMG(611)),
    (9,  "Yves Saint Laurent","Libre EDP",                      "Mujer",   90, "Lavanda, azahar, vainilla madagascar. Floral sensual moderno.",             9200, None,  "SI", "NO", IMG(56077)),
    (10, "Carolina Herrera",  "212 VIP Rosé EDP",               "Mujer",   80, "Champagne rosé, frutos rojos, musgo blanco. Festivo y femenino.",           5850, 6500,  "SI", "SI", IMG(22857)),
    (51, "Chanel",            "Chance EDT",                     "Mujer",  100, "Pimienta rosa, jazmín, pachulí, vetiver. Fresco sofisticado.",             10800, None, "SI", "NO", IMG(610)),
    (52, "Chanel",            "Chance Eau Tendre EDT",          "Mujer",  100, "Pomelo, membrillo, jazmín, iris. Frutal floral luminoso.",                 10500, None, "SI", "NO", IMG(8069)),
    (53, "Chanel",            "Chance Eau de Parfum",           "Mujer",  100, "Pimienta, jazmín, rosa, pachulí, amber. Más denso y oriental.",            11500, None, "SI", "NO", IMG(31351)),
    (54, "Chanel",            "No. 5 EDP",                      "Mujer",  100, "Aldehídos, ylang-ylang, rosa mayo, jazmín, vainilla, sándalo. Legendario.", 12500, None, "SI", "NO", IMG(40069)),
    (55, "Chanel",            "Coco Noir EDP",                  "Mujer",  100, "Bergamota, rosa, jazmín, pachulí, tonka. Oriental profundo nocturno.",     11000, None, "SI", "NO", IMG(15963)),
    (56, "Lancôme",           "Idôle EDP",                      "Mujer",   75, "Pera, bergamota, rosa, jazmín, almizcle blanco. Floral chypre ligero.",     8500, None, "SI", "NO", IMG(55795)),
    (57, "Lancôme",           "Miracle EDP",                    "Mujer",  100, "Jengibre, pimienta, lychee, magnolia, jazmín, almizcle. Fresco floral.",    7800, None, "SI", "NO", IMG(184)),
    (58, "Lancôme",           "La Vie Est Belle Intensément",   "Mujer",  100, "Iris, pera, jazmín, vainilla, pachulí. Más intenso que el original.",       9800, None, "SI", "NO", IMG(59326)),
    (59, "Carolina Herrera",  "Good Girl Supreme EDP",          "Mujer",   80, "Frutos rojos, jazmín egipcio, tonka, vetiver. Más gourmet y oscuro.",       7500, None, "SI", "NO", IMG(61769)),
    (60, "Carolina Herrera",  "Very Good Girl EDP",             "Mujer",   80, "Grosella negra, rosa, jazmín, cacao. Rojo sensual, gourmand floral.",       7800, None, "SI", "NO", IMG(65560)),
    (61, "Carolina Herrera",  "CH Women EDP",                   "Mujer",  100, "Pomelo, bergamota, rosa, ámbar, cuero. Elegante clásico.",                  5800, None, "SI", "NO", ""),
    (62, "Carolina Herrera",  "212 EDT",              "Mujer",  100, "Mandarina, flor de loto, bambú, sándalo. Fresco urbano limpio.",            5500, None, "SI", "NO", IMG(296)),
    (63, "Carolina Herrera",  "212 Sexy EDP",                   "Mujer",  100, "Pimienta rosa, gardenia, vainilla, musgo. Oriental floral sensual.",        5800, None, "SI", "NO", IMG(306)),
    (64, "Yves Saint Laurent","Libre Intense EDP",              "Mujer",   90, "Lavanda, azahar, vainilla, ámbar negro. Más oscuro y nocturno.",           10200, None, "SI", "NO", IMG(62318)),
    (65, "Yves Saint Laurent","Black Opium EDP",                "Mujer",   90, "Café, jazmín sambac, vainilla, praliné. Oriental gourmand adictivo.",       9500, None, "SI", "NO", IMG(25324)),
    (66, "Yves Saint Laurent","Mon Paris EDP",                  "Mujer",   90, "Fresa, pera, peonía, jazmín, pachulí. Floral frutal chypre moderno.",       9200, None, "SI", "NO", IMG(38914)),
    (67, "Dior",              "J'adore EDP",                    "Mujer",  100, "Pera, magnolia, jazmín sambac, rosa, vainilla. Floral icónico luminoso.",  12500, None, "SI", "NO", IMG(210)),
    (68, "Dior",              "Miss Dior EDP",                  "Mujer",  100, "Rosa grasse, lily of the valley, pachulí. Floral romántico moderno.",      11500, None, "SI", "NO", IMG(45202)),
    (69, "Dior",              "Poison Girl EDP",                "Mujer",  100, "Pomelo, rosa damascena, almendra amarga, vainilla. Oriental frutal.",      10500, None, "SI", "NO", IMG(35561)),
    (70, "Giorgio Armani",    "Sì EDP",                         "Mujer",  100, "Cassis, rosa de mayo, vainilla, pachulí. Chypre moderno frutal.",          10200, None, "SI", "NO", IMG(18453)),
    (71, "Giorgio Armani",    "My Way EDP",                     "Mujer",   90, "Bergamota, azahar, tuberosa, vainilla, cedro blanco. Floral moderno.",      9200, None, "SI", "NO", IMG(62036)),
    (72, "Giorgio Armani",    "Sì Passione EDP",                "Mujer",  100, "Pera, cassis, rosa, frangipani, vainilla. Más dulce y pasional que Sì.",    9800, None, "SI", "NO", IMG(48002)),
    (73, "Paco Rabanne",      "Lady Million EDP",               "Mujer",   80, "Frambuesa, neroli, jazmín, naranjo, miel, pachulí. Dorado y divertido.",    6500, None, "SI", "NO", IMG(9045)),
    (74, "Paco Rabanne",      "Olympéa EDP",                    "Mujer",   80, "Jazmín acuático, mandarina verde, vainilla salada, sándalo. Solar sexy.",   6200, None, "SI", "NO", IMG(31666)),
    (75, "Paco Rabanne",      "Fame EDP",                       "Mujer",   80, "Mango, bergamota, jazmín, olíbano, vainilla, sándalo. Mineral floral.",     6800, None, "SI", "NO", IMG(74962)),
    (76, "Calvin Klein",      "Euphoria EDP",                   "Mujer",  100, "Pomelo, granada, flor loto, orquídea, amber, cream accord. Oriental seductor.", 6500, None, "SI", "NO", IMG(253)),
    (77, "Calvin Klein",      "Eternity Women EDP",             "Mujer",  100, "Freesia, clavel, madreselva, sándalo. Floral clásico.",                     5800, None, "SI", "NO", IMG(257)),
    (78, "Versace",           "Bright Crystal EDT",             "Mujer",   90, "Yuzu, granada, magnolia, peonía, lotus, almizcle. Floral acuático.",        6200, None, "SI", "NO", IMG(632)),
    (79, "Versace",           "Eros Pour Femme EDP",            "Mujer",  100, "Limón, mandarina, peonía, jazmín, ámbar, sándalo. Floral oriental.",        6800, None, "SI", "NO", IMG(28958)),
    (80, "Hugo Boss",         "Boss Alive EDP",                 "Mujer",   80, "Pera, ciruela, jazmín, vainilla, cedro. Floral moderno femenino.",          5800, None, "SI", "NO", IMG(59228)),
    (81, "Hugo Boss",         "Boss Ma Vie EDP",                "Mujer",   75, "Cactus, rosa, jazmín, sándalo. Verde floral fresco.",                       5500, None, "SI", "NO", IMG(25298)),
    (82, "Tommy Hilfiger",    "Tommy Girl EDT",                 "Mujer",  100, "Manzana, camelia, rosa, lirio, bayas silvestres. Americana fresca.",        4800, None, "SI", "NO", IMG(3016)),
    (83, "Jean Paul Gaultier","La Belle EDP",                   "Mujer",  100, "Pera, vainilla bourbon, habas tonka, flor de sal. Gourmand luminoso.",      7500, None, "SI", "NO", IMG(55786)),
    (84, "Jean Paul Gaultier","Classique EDT",                  "Mujer",  100, "Anís, mandarina, rosa, jengibre, vainilla, ámbar. Oriental icónico.",       7200, None, "SI", "NO", IMG(394)),
    (85, "Kenzo",             "Flower by Kenzo EDP",            "Mujer",  100, "Mandarina, grosella negra, violeta, rosa, haba tonka, almizcle blanco.",    6800, None, "SI", "NO", IMG(72)),
    (86, "Nina Ricci",        "Nina EDT",                       "Mujer",   80, "Limón, lima, manzana dulce, peonía, vainilla. Frutal floral gourmand.",     5500, None, "SI", "NO", IMG(147)),
    (87, "Nina Ricci",        "Luna EDT",                       "Mujer",   80, "Pomelo, caramelo, azahar, algodón de azúcar. Frutal dulce joven.",          5200, None, "SI", "NO", IMG(39376)),
    (88, "Bvlgari",           "Omnia Crystalline EDT",          "Mujer",   65, "Bambú, nashi, loto, teca, madreperla. Acuático transparente.",              5800, None, "SI", "NO", IMG(152)),
    (89, "Chloé",             "Chloé EDP",                      "Mujer",   75, "Peonía, lichi, freesia, rosa, magnolia, cedro. Floral rosado icónico.",     9500, None, "SI", "NO", IMG(1733)),
    (90, "Bvlgari",           "Aqva Pour Homme EDT",            "Hombre", 100, "Mandarina, petitgrain, almíbar, musgo marino. Fresco acuático mineral.",    6800, None, "SI", "NO", IMG(153)),
]

# ───── Escribir al xlsx ─────
ws.append(headers)
for col_idx, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = Font(name="Arial", bold=True, color=BLANCO, size=11)
    c.fill = PatternFill("solid", start_color=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(bottom=Side(style="medium", color=AZUL))

for row in data:
    ws.append(row)

thin = Side(style="thin", color="D1D5DB")
for row_idx in range(2, 2 + len(data)):
    fill = BLANCO if row_idx % 2 == 0 else AZUL_CLARO
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=fill)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(row=row_idx, column=7).number_format = '"$"#,##0'
    ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row_idx, column=8).number_format = '"$"#,##0;;-'
    ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for c_idx in (4, 5, 9, 10):
        ws.cell(row=row_idx, column=c_idx).alignment = Alignment(horizontal="center", vertical="center")
    if ws.cell(row=row_idx, column=10).value == "SI":
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", start_color="FDF4E0")

widths = {"A": 6, "B": 20, "C": 34, "D": 10, "E": 10, "F": 55,
          "G": 14, "H": 16, "I": 10, "J": 12, "K": 55}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.row_dimensions[1].height = 30
for r in range(2, 2 + len(data)):
    ws.row_dimensions[r].height = 40

dv_si_no = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
ws.add_data_validation(dv_si_no)
dv_si_no.add(f"I2:I{1 + len(data)}")
dv_promo = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
ws.add_data_validation(dv_promo)
dv_promo.add(f"J2:J{1 + len(data)}")
dv_gen = DataValidation(type="list", formula1='"Hombre,Mujer,Unisex"', allow_blank=False)
ws.add_data_validation(dv_gen)
dv_gen.add(f"D2:D{1 + len(data)}")

ws.freeze_panes = "A2"

# ───── Hoja LEEME ─────
ws2 = wb.create_sheet("LEEME")
inst = [
    ("Catálogo Salvador Perfumería — 90 perfumes", ""),
    ("", ""),
    ("1. Importá este archivo a Google Sheets:", "Drive → Nuevo → Subir archivo → Abrir con Google Sheets"),
    ("2. Compartí público:", "Compartir → 'Cualquier usuario con el vínculo' → Lector"),
    ("3. Publicá como CSV:", "Archivo → Compartir → Publicar en la web → formato CSV"),
    ("4. Pegá el link en index.html (variable sheetCsvUrl).", ""),
    ("", ""),
    ("Columnas (no renombrar ni mover):", ""),
    ("  Stock = SI/NO (visible en el catálogo)", ""),
    ("  Promocion = SI/NO (aparece con badge dorado y borde especial)", ""),
    ("  Precio antes (UYU) = precio tachado (opcional, solo para promos)", ""),
    ("  Imagen URL = URL pública de la foto (Fragrantica CDN o subida propia)", ""),
    ("", ""),
    ("Notas:", ""),
    ("  • Los precios son benchmark Farmashop abr 2026. Revisá y ajustá.", ""),
    ("  • Las imágenes vienen de Fragrantica CDN. Si alguna no carga, reemplazá la URL.", ""),
    ("  • Podés agregar/quitar perfumes editando el Google Sheet, no hace falta tocar código.", ""),
]
for row in inst:
    ws2.append(row)
ws2["A1"].font = Font(name="Arial", bold=True, color=BLANCO, size=14)
ws2["A1"].fill = PatternFill("solid", start_color=AZUL)
ws2.merge_cells("A1:B1")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28
for r in range(3, len(inst) + 1):
    ws2.cell(row=r, column=1).font = Font(name="Arial", size=10, bold=True, color=AZUL)
    ws2.cell(row=r, column=2).font = Font(name="Arial", size=10, color=GRIS)
    ws2.cell(row=r, column=1).alignment = Alignment(vertical="center", wrap_text=True)
    ws2.cell(row=r, column=2).alignment = Alignment(vertical="center", wrap_text=True)
ws2.column_dimensions["A"].width = 55
ws2.column_dimensions["B"].width = 75

import os, json
root = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.join(root, "data", "catalogo.xlsx")
os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
wb.save(xlsx_path)
print(f"✓ Saved xlsx: {xlsx_path} ({len(data)} perfumes)")

# ───── Generar también el JSON para inyectar en index.html seed ─────
def yes(v): return v == "SI"
perfumes_json = []
for row in data:
    id_, marca, nombre, genero, ml, notas, precio, precio_antes, stock, promocion, imagen = row
    perfumes_json.append({
        "id": id_, "marca": marca, "nombre": nombre, "genero": genero,
        "ml": ml, "notas": notas, "precio": precio,
        "precio_antes": precio_antes if precio_antes else None,
        "stock": yes(stock), "promocion": yes(promocion),
        "imagen": imagen,
    })
# Orden final por ID para que aparezcan prolijos
perfumes_json.sort(key=lambda p: p["id"])
json_str = json.dumps(perfumes_json, ensure_ascii=False, separators=(',', ':'))

# Inyectar el JSON en el seed de index.html
index_path = os.path.join(root, "index.html")
with open(index_path, "r", encoding="utf-8") as f:
    html = f.read()
import re
new_html = re.sub(
    r'(<script id="seed-data" type="application/json">)[\s\S]*?(</script>)',
    lambda m: m.group(1) + "\n" + json_str + "\n" + m.group(2),
    html,
    count=1,
)
with open(index_path, "w", encoding="utf-8") as f:
    f.write(new_html)
print(f"✓ Updated seed in index.html ({len(perfumes_json)} perfumes)")

# Resumen
gen = {}
promos = 0
with_img = 0
for p in perfumes_json:
    gen[p["genero"]] = gen.get(p["genero"], 0) + 1
    if p["promocion"]: promos += 1
    if p["imagen"]: with_img += 1
print(f"  Distribución: {gen}")
print(f"  Promociones: {promos}")
print(f"  Con imagen: {with_img}/{len(perfumes_json)}")
